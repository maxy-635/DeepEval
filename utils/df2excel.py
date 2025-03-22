import os
import openpyxl


class DataFrame2Excel:
    """
    Save DataFrame data to excel file
    :param data: DataFrame data
    :param excel_path: excel file path
    :return
    """

    def __init__(self, data, save_excel_path):
        self.data = data
        self.save_excel_path = save_excel_path

    def df2excel(self, sheet_name):
        if os.path.exists(self.save_excel_path):
            wb = openpyxl.load_workbook(self.save_excel_path)
        else:
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)  

        sheet = wb.create_sheet(title=sheet_name)
        for col_idx, col_name in enumerate(self.data.columns, start=1):
            sheet.cell(row=1, column=col_idx, value=col_name)
        for r_idx, row in enumerate(self.data.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        if not os.path.exists(os.path.dirname(self.save_excel_path)):
            os.makedirs(os.path.dirname(self.save_excel_path))
        wb.save(self.save_excel_path)
